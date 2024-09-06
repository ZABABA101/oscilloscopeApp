from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries

import pyvisa as visa
from pyscope import *
from measurements import *
from utilities import *

def insertNum(ws, numbers, cells):
    """
    插入数字到指定的单元格中。
    参数:
    ws: 工作表对象
    numbers: 数字列表
    cells: 单元格位置列表
    """
    for number, cell in zip(numbers, cells):
        if cell in ws.merged_cells:  # 检查单元格是否为合并单元格
            cell_range = ws.merged_cells.ranges
            for range_ in cell_range:
                if cell in range_.coord:
                    # 仅在合并单元格的起始位置写入数据
                    start_cell = range_.start_cell.coordinate
                    ws[start_cell].value = number
                    break
        else:
            ws[cell].value = number

def insertPng(ws, merged_cell_address, image_path):
    """
    插入PNG图片到指定的合并单元格中。
    参数:
    ws: 工作表对象
    merged_cell_address: 合并单元格的起始地址
    image_path: 图片的文件路径
    """
    img = Image(image_path)
    # 找到合并单元格的范围和尺寸
    for range_ in ws.merged_cells.ranges:
        if merged_cell_address in range_.coord:
            min_col, min_row, max_col, max_row = range_boundaries(range_.coord)
            width = sum((ws.column_dimensions[chr(64 + col)].width or 8.43 for col in range(min_col, max_col + 1)))
            height = sum((ws.row_dimensions[row].height or 15 for row in range(min_row, max_row + 1)))
            img.width = width * 6  # 调整宽度
            img.height = height * 0.75  # 调整高度
            img.anchor = merged_cell_address
            ws.add_image(img)
            break

#--------------------------------------------------------------------------------
#  获取波形和数据
#---------------------------------------------------------------------------------
rm = visa.ResourceManager()
print(rm.list_resources())
scope_address = 'USB0::0x0699::0x0530::C069124::INSTR'
scope = TekScopeMainstream(rm, scope_address, 10000)

channels = ["CH1","CH2","CH3","CH4"]
vertScales = [1, 1, 1, 1]
vertPositions = [0, 0, 0, 0]
vertOffsets = [0, 0, 0, 0]
scope.display_analog_channel(channels, [1, 1, 1, 1])
scope.set_vertical(channels, vertScales, vertOffsets, vertPositions)
scope.set_horizontal(1.25e9, 12.5e6)
record_len = scope.get_record_length()
#获取截图信息
scope.transfer_screenshot('E:/ScopeApplication/ScopeFile','8_22.png')
#获取数据信息
scope.acquire_run_single_auto_acq_complete()
volt_data, time_data = scope.transfer_wfm(channels, 1, record_len)
base, top = calc_base_top(volt_data[0])
logger.info(f"the base voltage is {base:.5f} V, the top voltage is {top:.5f} V")

crossTypes, crossTimes = calc_crosses(time_data, volt_data[0], 0.5, 0.2)

freq_mean, freq_max, freq_min, freq_var = calc_pulse_frequency(crossTypes, crossTimes)
logger.info(f"the pulse frequency is {freq_mean:.5e} @mean, {freq_max:.5e} @max, {freq_min:.5e} @min, {freq_var:.5e} @var")
#--------------------------------------------------------------------------------
#获取完波形和数据进行excel的编辑操作
#---------------------------------------------------------------------------------
file_path = 'E:/ScopeApplication/ReportExcel/report.xlsx'
image_path = 'E:/ScopeApplication/ScopeFile/8_22.png'
merged_cell_address = 'A2'
numbers = []
numbers.append(freq_max)
numbers.append(freq_min)
numbers.append(freq_mean)
numbers.append(freq_var)

print(numbers)

cells = ["D2", "F2", "H2", "J2"]

wb = load_workbook(file_path)
ws = wb.active


# 调用函数插入图片和数字
insertPng(ws, merged_cell_address, image_path)
insertNum(ws, numbers, cells)

# 保存 Excel 文件
wb.save(file_path)
