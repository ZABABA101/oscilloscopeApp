#----------------------------------------------------
# author:zhangao
# date:2024.8.22
# description : Base Class to Excute Some Opetation of Excel Sheet
#-----------------------------------------------------
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries
class ExcelBase(object):

    def insertNum(ws, numbers, cells):
        """
        插入数字到指定的单元格中。
        参数:
        ws: 工作表对象
        numbers: 数字列表
        cells: 单元格位置列表
        """
        for number, cell in zip(numbers, cells):
            if cell in ws.merged_cells:  # 检查单元格是否为合并单元格
                cell_range = ws.merged_cells.ranges
                for range_ in cell_range:
                    if cell in range_.coord:
                        # 仅在合并单元格的起始位置写入数据
                        start_cell = range_.start_cell.coordinate
                        ws[start_cell].value = number
                        break
            else:
                ws[cell].value = number


def insertPng(ws, merged_cell_address, image_path):
    """
    插入PNG图片到指定的合并单元格中。
    参数:
    ws: 工作表对象
    merged_cell_address: 合并单元格的起始地址
    image_path: 图片的文件路径
    """
    img = Image(image_path)
    # 找到合并单元格的范围和尺寸
    for range_ in ws.merged_cells.ranges:
        if merged_cell_address in range_.coord:
            min_col, min_row, max_col, max_row = range_boundaries(range_.coord)
            width = sum((ws.column_dimensions[chr(64 + col)].width or 8.43 for col in range(min_col, max_col + 1)))
            height = sum((ws.row_dimensions[row].height or 15 for row in range(min_row, max_row + 1)))
            img.width = width * 6  # 调整宽度
            img.height = height * 0.75  # 调整高度
            img.anchor = merged_cell_address
            ws.add_image(img)
            break